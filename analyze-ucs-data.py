#!/usr/bin/env python3
"""
Analyze UCS Satellite Database Excel files
Shows column structure, data samples, and assesses value for merging
"""

import openpyxl
import json
from pathlib import Path
from collections import defaultdict

def analyze_xlsx(file_path):
    """Analyze an Excel file and return detailed info"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {Path(file_path).name}")
    print(f"{'='*80}")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Dimensions: {sheet.dimensions}")

        # Get headers (first row)
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)

        print(f"\nColumns ({len(headers)}):")
        for i, header in enumerate(headers, 1):
            print(f"  {i:2}. {header}")

        # Count rows
        row_count = sheet.max_row - 1  # Exclude header
        print(f"\nTotal rows: {row_count:,}")

        # Sample data from first few rows (rows 2-6)
        print(f"\nSample data (first 5 records):")
        sample_data = []
        for row_num in range(2, min(7, sheet.max_row + 1)):
            row_data = {}
            for col_num, header in enumerate(headers, 1):
                cell_value = sheet.cell(row_num, col_num).value
                row_data[header] = cell_value
            sample_data.append(row_data)

        # Print sample in readable format
        for i, record in enumerate(sample_data, 1):
            print(f"\n  Record {i}:")
            for key, value in record.items():
                if value is not None and value != '':
                    print(f"    {key}: {value}")

        # Analyze data distribution for key fields
        print(f"\n\nData Analysis:")

        # Check for NORAD Cat IDs
        if 'NORAD Number' in headers:
            col_idx = headers.index('NORAD Number') + 1
            norad_ids = set()
            for row_num in range(2, sheet.max_row + 1):
                val = sheet.cell(row_num, col_idx).value
                if val:
                    norad_ids.add(val)
            print(f"  Unique NORAD IDs: {len(norad_ids):,}")

        # Check for countries/operators
        if 'Country/Org of UN Registry' in headers:
            col_idx = headers.index('Country/Org of UN Registry') + 1
            countries = defaultdict(int)
            for row_num in range(2, sheet.max_row + 1):
                val = sheet.cell(row_num, col_idx).value
                if val:
                    countries[val] += 1
            print(f"  Countries/Orgs: {len(countries)}")
            print(f"  Top 5:")
            for country, count in sorted(countries.items(), key=lambda x: x[1], reverse=True)[:5]:
                print(f"    {country}: {count}")

        # Check for purposes/applications
        if 'Purpose' in headers:
            col_idx = headers.index('Purpose') + 1
            purposes = defaultdict(int)
            for row_num in range(2, sheet.max_row + 1):
                val = sheet.cell(row_num, col_idx).value
                if val:
                    purposes[val] += 1
            print(f"  Purposes: {len(purposes)}")
            print(f"  Top 5:")
            for purpose, count in sorted(purposes.items(), key=lambda x: x[1], reverse=True)[:5]:
                print(f"    {purpose}: {count}")

        # Check for orbit classes
        if 'Class of Orbit' in headers:
            col_idx = headers.index('Class of Orbit') + 1
            orbits = defaultdict(int)
            for row_num in range(2, sheet.max_row + 1):
                val = sheet.cell(row_num, col_idx).value
                if val:
                    orbits[val] += 1
            print(f"  Orbit Classes:")
            for orbit, count in sorted(orbits.items(), key=lambda x: x[1], reverse=True):
                print(f"    {orbit}: {count}")

# Analyze both files
files = [
    '/home/major/aetherlink/docs/UCS-Satellite-Database 5-1-2023.xlsx',
    '/home/major/aetherlink/docs/UCS-Satellite-Database-Officialname 5-1-2023.xlsx'
]

for file in files:
    analyze_xlsx(file)

print(f"\n\n{'='*80}")
print("COMPARISON WITH EXISTING DATABASE")
print(f"{'='*80}")
print("""
Current database sources:
  - Space-Track.org: 31,660+ satellites with TLEs, orbital data
  - SatNOGS: Transmitter/frequency data for active satellites

UCS Database characteristics (as of May 2023):
  - More curated/selective dataset (typically 3,000-5,000 satellites)
  - Focused on operational satellites
  - Rich metadata: purpose, detailed ownership, applications
  - Physical characteristics: mass, power, lifetime
  - Launch details: vehicle, site
  - Contractor information

RECOMMENDATION:
""")

# Read one file to get actual column count
wb = openpyxl.load_workbook('/home/major/aetherlink/docs/UCS-Satellite-Database 5-1-2023.xlsx', data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
row_count = ws.max_row - 1

print(f"✓ MERGE - High value fields found:")
print(f"  Total UCS records: {row_count:,}")
print(f"  Total fields: {len(headers)}")
print(f"""
High-value fields to merge:
  1. Purpose/Application - categorizes satellite missions
  2. Detailed operator/ownership - more granular than SATCAT
  3. Launch mass, power, expected lifetime - physical specs
  4. Launch vehicle & launch site - launch details
  5. Contractor info - manufacturer data
  6. Users (government, commercial, civil, military)

These complement Space-Track (orbital/TLE) and SatNOGS (frequency) data.

MERGE STRATEGY:
  - Use NORAD ID as primary key for matching
  - Enrich existing satellite records with UCS metadata
  - Create new tables for UCS-specific fields
  - Handle data age: UCS is ~2.5 years old (May 2023)
    * Good for: historical satellites, launch details, specs
    * May be outdated for: operational status, current purpose
""")
