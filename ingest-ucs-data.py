#!/usr/bin/env python3
"""
Ingest UCS Satellite Database into PostgreSQL

This script enriches the satellite database with UCS metadata while
keeping Space-Track.org data as the authoritative source.

IMPORTANT: Space-Track data is AUTHORITATIVE for:
  - Orbital parameters (TLEs, epoch, inclination, etc.)
  - Current operational status
  - Decay dates

UCS data provides ENRICHMENT for:
  - Mission purpose and applications
  - Physical characteristics (mass, power)
  - Launch details (vehicle, site, contractor)
  - Detailed ownership information
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from datetime import datetime
import os
import sys

# Database connection (from satcat-backend .env)
DB_CONFIG = {
    'dbname': os.environ.get('DB_NAME', 'satcat_db'),
    'user': os.environ.get('DB_USER', 'major'),
    'password': os.environ.get('DB_PASSWORD', '1234'),
    'host': os.environ.get('DB_HOST', 'localhost'),
    'port': os.environ.get('DB_PORT', '5432')
}

# UCS Excel file path (use the main file with alternate names)
UCS_FILE = '/home/major/aetherlink/docs/UCS-Satellite-Database 5-1-2023.xlsx'

def clean_string(value):
    """Clean and normalize string values"""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value and value != 'NR' else None
    return str(value)

def clean_number(value):
    """Clean and convert to number, handling 0 as None for some fields"""
    if value is None or value == '' or value == 'NR':
        return None
    try:
        num = float(value)
        return num if num != 0 else None
    except (ValueError, TypeError):
        return None

def parse_ucs_data():
    """Parse UCS Excel file and return list of satellite records"""
    print(f"Opening UCS database: {UCS_FILE}")
    workbook = openpyxl.load_workbook(UCS_FILE, data_only=True)
    sheet = workbook.active

    # Get headers
    headers = [cell.value for cell in sheet[1]]

    # Map column names to indices
    col_map = {header: idx for idx, header in enumerate(headers, 1)}

    print(f"Found {sheet.max_row - 1:,} satellite records")
    print("Parsing UCS data...")

    satellites = []
    skipped = 0

    for row_num in range(2, sheet.max_row + 1):
        try:
            # Get NORAD ID (required for matching)
            norad_id = sheet.cell(row_num, col_map.get('NORAD Number')).value

            if not norad_id:
                skipped += 1
                continue

            norad_id = int(norad_id)

            # Extract UCS metadata
            sat_data = {
                'norad_id': norad_id,
                'purpose': clean_string(sheet.cell(row_num, col_map.get('Purpose')).value),
                'detailed_purpose': clean_string(sheet.cell(row_num, col_map.get('Detailed Purpose')).value),
                'users': clean_string(sheet.cell(row_num, col_map.get('Users')).value),

                'country_of_operator': clean_string(sheet.cell(row_num, col_map.get('Country of Operator/Owner')).value),
                'operator_owner': clean_string(sheet.cell(row_num, col_map.get('Operator/Owner')).value),
                'country_un_registry': clean_string(sheet.cell(row_num, col_map.get('Country/Org of UN Registry')).value),

                'launch_mass_kg': clean_number(sheet.cell(row_num, col_map.get('Launch Mass (kg.)')).value),
                'dry_mass_kg': clean_number(sheet.cell(row_num, col_map.get('Dry Mass (kg.)')).value),
                'power_watts': clean_number(sheet.cell(row_num, col_map.get('Power (watts)')).value),
                'expected_lifetime_yrs': clean_number(sheet.cell(row_num, col_map.get('Expected Lifetime (yrs.)')).value),

                'launch_vehicle': clean_string(sheet.cell(row_num, col_map.get('Launch Vehicle')).value),
                'launch_site': clean_string(sheet.cell(row_num, col_map.get('Launch Site')).value),
                'contractor': clean_string(sheet.cell(row_num, col_map.get('Contractor')).value),
                'country_of_contractor': clean_string(sheet.cell(row_num, col_map.get('Country of Contractor')).value),

                'ucs_orbit_class': clean_string(sheet.cell(row_num, col_map.get('Class of Orbit')).value),
                'ucs_orbit_type': clean_string(sheet.cell(row_num, col_map.get('Type of Orbit')).value),
                'ucs_longitude_geo': clean_number(sheet.cell(row_num, col_map.get('Longitude of GEO (degrees)')).value),

                'cospar_number': clean_string(sheet.cell(row_num, col_map.get('COSPAR Number')).value),
                'comments': clean_string(sheet.cell(row_num, col_map.get('Comments')).value),
                'alternate_names': clean_string(sheet.cell(row_num, col_map.get('Name of Satellite, Alternate Names')).value),
            }

            satellites.append(sat_data)

            if row_num % 1000 == 0:
                print(f"  Processed {row_num - 1:,} rows...")

        except Exception as e:
            print(f"  Warning: Error processing row {row_num}: {e}")
            skipped += 1
            continue

    print(f"Parsed {len(satellites):,} satellite records")
    if skipped > 0:
        print(f"Skipped {skipped} records (missing NORAD ID or errors)")

    return satellites

def ingest_to_database(satellites):
    """Insert UCS metadata into PostgreSQL database"""
    print("\nConnecting to database...")

    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()

    try:
        # First, run the migration to create the table
        print("Creating/verifying ucs_satellite_metadata table...")

        migration_file = '/home/major/aetherlink/satcat-backend/prisma/migrations/20251109_add_ucs_metadata/migration.sql'
        with open(migration_file, 'r') as f:
            migration_sql = f.read()
            cursor.execute(migration_sql)
            conn.commit()

        print("Table ready!")

        # Check how many satellites exist in main table
        cursor.execute("SELECT COUNT(*) FROM satellite")
        total_sats = cursor.fetchone()[0]
        print(f"Main satellite table has {total_sats:,} records from Space-Track")

        # Get all existing NORAD IDs from Space-Track
        print("Fetching existing NORAD IDs...")
        cursor.execute("SELECT norad_id FROM satellite")
        existing_norads = set(row[0] for row in cursor.fetchall())
        print(f"Found {len(existing_norads):,} unique NORAD IDs in Space-Track data")

        # Deduplicate UCS data by NORAD ID (keep first occurrence)
        seen_norads = set()
        unique_satellites = []
        duplicates = 0
        for s in satellites:
            if s['norad_id'] not in seen_norads:
                seen_norads.add(s['norad_id'])
                unique_satellites.append(s)
            else:
                duplicates += 1
        
        if duplicates > 0:
            print(f"Removed {duplicates} duplicate NORAD IDs from UCS data")
        satellites = unique_satellites

        # Filter UCS data to only include satellites that exist in Space-Track
        matched_satellites = [s for s in satellites if s['norad_id'] in existing_norads]
        unmatched_count = len(satellites) - len(matched_satellites)

        print(f"\nMatching UCS data with Space-Track:")
        print(f"  UCS satellites matching Space-Track: {len(matched_satellites):,}")
        print(f"  UCS satellites NOT in Space-Track:   {unmatched_count:,} (will be skipped)")

        if len(matched_satellites) == 0:
            print("\nNo matching satellites found! Exiting.")
            return

        # Prepare insert query
        insert_query = """
            INSERT INTO ucs_satellite_metadata (
                norad_id, purpose, detailed_purpose, users,
                country_of_operator, operator_owner, country_un_registry,
                launch_mass_kg, dry_mass_kg, power_watts, expected_lifetime_yrs,
                launch_vehicle, launch_site, contractor, country_of_contractor,
                ucs_orbit_class, ucs_orbit_type, ucs_longitude_geo,
                cospar_number, comments, alternate_names
            ) VALUES %s
            ON CONFLICT (norad_id)
            DO UPDATE SET
                purpose = EXCLUDED.purpose,
                detailed_purpose = EXCLUDED.detailed_purpose,
                users = EXCLUDED.users,
                country_of_operator = EXCLUDED.country_of_operator,
                operator_owner = EXCLUDED.operator_owner,
                country_un_registry = EXCLUDED.country_un_registry,
                launch_mass_kg = EXCLUDED.launch_mass_kg,
                dry_mass_kg = EXCLUDED.dry_mass_kg,
                power_watts = EXCLUDED.power_watts,
                expected_lifetime_yrs = EXCLUDED.expected_lifetime_yrs,
                launch_vehicle = EXCLUDED.launch_vehicle,
                launch_site = EXCLUDED.launch_site,
                contractor = EXCLUDED.contractor,
                country_of_contractor = EXCLUDED.country_of_contractor,
                ucs_orbit_class = EXCLUDED.ucs_orbit_class,
                ucs_orbit_type = EXCLUDED.ucs_orbit_type,
                ucs_longitude_geo = EXCLUDED.ucs_longitude_geo,
                cospar_number = EXCLUDED.cospar_number,
                comments = EXCLUDED.comments,
                alternate_names = EXCLUDED.alternate_names,
                updated_at = CURRENT_TIMESTAMP
        """

        # Convert to tuples for bulk insert
        values = [
            (
                s['norad_id'], s['purpose'], s['detailed_purpose'], s['users'],
                s['country_of_operator'], s['operator_owner'], s['country_un_registry'],
                s['launch_mass_kg'], s['dry_mass_kg'], s['power_watts'], s['expected_lifetime_yrs'],
                s['launch_vehicle'], s['launch_site'], s['contractor'], s['country_of_contractor'],
                s['ucs_orbit_class'], s['ucs_orbit_type'], s['ucs_longitude_geo'],
                s['cospar_number'], s['comments'], s['alternate_names']
            )
            for s in matched_satellites
        ]

        print(f"\nInserting {len(values):,} UCS metadata records...")
        print("(Using ON CONFLICT to update existing records)")

        # Bulk insert
        execute_values(cursor, insert_query, values, page_size=1000)
        conn.commit()

        # Get statistics
        cursor.execute("""
            SELECT
                COUNT(*) as total_ucs,
                COUNT(s.norad_id) as matched_spacetrack
            FROM ucs_satellite_metadata u
            LEFT JOIN satellite s ON u.norad_id = s.norad_id
        """)

        total_ucs, matched = cursor.fetchone()

        print(f"\n{'='*70}")
        print("INGESTION COMPLETE!")
        print(f"{'='*70}")
        print(f"UCS records inserted:        {total_ucs:,}")
        print(f"Matched with Space-Track:    {matched:,} (100%)")
        print(f"UCS-only records skipped:    {unmatched_count:,}")
        print(f"\nSpace-Track remains authoritative for orbital data.")
        print(f"UCS provides enrichment: purpose, specs, launch details.")

        # Show purpose breakdown
        cursor.execute("""
            SELECT purpose, COUNT(*) as count
            FROM ucs_satellite_metadata
            WHERE purpose IS NOT NULL
            GROUP BY purpose
            ORDER BY count DESC
            LIMIT 10
        """)

        print(f"\nTop 10 satellite purposes:")
        for purpose, count in cursor.fetchall():
            print(f"  {purpose:30} {count:>6,}")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        cursor.close()
        conn.close()

def main():
    print("="*70)
    print("UCS Satellite Database Ingestion")
    print("="*70)
    print()
    print("AUTHORITY HIERARCHY:")
    print("  1. Space-Track.org  = AUTHORITATIVE for orbital data")
    print("  2. UCS Database     = ENRICHMENT for mission/specs data")
    print()

    # Parse UCS data
    satellites = parse_ucs_data()

    # Ingest to database
    ingest_to_database(satellites)

    print(f"\n{'='*70}")
    print("SUCCESS!")
    print(f"{'='*70}")
    print("\nYour satellite database now includes UCS enrichment data.")
    print("Query the 'ucs_satellite_metadata' table for mission details.")
    print("\nSpace-Track data remains unchanged and authoritative.")

if __name__ == '__main__':
    main()
# Patch: Add this after line 180 in the ingest_to_database function, before filtering
